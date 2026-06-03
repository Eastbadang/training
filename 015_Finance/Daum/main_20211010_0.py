# xlsx 파일 불러드리고, 시트 및 시트 내용 확인

from openpyxl import load_workbook

wb = load_workbook('D:/Works/PycharmProjects/001_TRAINING/015_Finance/Daum/daum_finance_20211010.xlsx')

for sheet_nm in wb.sheetnames:
    # print('*' * 100)
    print('시트명:', sheet_nm)
#    sheet = wb[sheet_nm]

#    for row_data in sheet.iter_rows(min_row=1):
#        for cell in row_data:
#            print('[', cell.value, ']')
#            # print('=' * 100)

sheet = wb['Sheet3']
for row_data in sheet.iter_rows(min_row=1):
    for cell in row_data:
        print('[', cell.value, ']')
        # print('=' * 100)

sheet = wb['현재가']
print(sheet.cell(row=1, column=4).value)

wb.close()