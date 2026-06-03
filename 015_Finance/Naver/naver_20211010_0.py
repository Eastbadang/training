# 1. 네이버 증권 기볼 설정 외 항목 결과
# 2. Flask 기초
# 3. 웹 상에서 크롤링한 데이터 보여주기

# 1. 라이브러리 호출
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup as bs

# 2. 데이터 요청
# 엑셀 파일 처리
wb = load_workbook('D:/Works/PycharmProjects/001_TRAINING/015_Finance/Daum/daum_finance_20211010.xlsx')
#for sheet_nm in wb.sheetnames:
    # print('*' * 100)
    # print('시트명:', sheet_nm)
#    sheet = wb[sheet_nm]

#    for row_data in sheet.iter_rows(min_row=1):
#        for cell in row_data:
#            print('[', cell.value, ']')
#            # print('=' * 100)

# 웹 크롤링 처리
base_url = "https://finance.naver.com/sise/field_submit.nhn?menu=market_sum&returnUrl=http://finance.naver.com/sise/sise_market_sum.nhn?sosok=0&fieldIds=quant&fieldIds=ask_buy&fieldIds=market_sum&fieldIds=per&fieldIds=ask_sell&fieldIds=roe"
req = requests.get(base_url)
html = req.text
soup = bs(html, 'lxml')

#print(soup)

# 3. 데이터 파싱
headers = soup.select("#contentarea > div.box_type_l > table.type_2 > thead > tr > th")
headers = list(map(lambda x: x.text, headers))
# print(headers)

sheet = wb['NAVER']
i = 1
sheet.cell(row=1, column=1).value = "순위"
sheet.cell(row=1, column=2).value = "종목명"
sheet.cell(row=1, column=3).value = "현재가"
sheet.cell(row=1, column=4).value = "매수호가"
sheet.cell(row=1, column=5).value = "매도호가"
sheet.cell(row=1, column=6).value = "PER"
sheet.cell(row=1, column=7).value = "ROE"

stock_datas = soup.select("#contentarea > div.box_type_l > table.type_2 > tbody > tr")  # 공통 주소
for data in stock_datas:
    try:
        i = i + 1
        stock_rank = data.select_one("td.no").text
        stock_name = data.select_one("td:nth-child(2) > a").text
        stock_price = data.select_one("td:nth-child(3)").text
        stock_quant = data.select_one("td:nth-child(7)").text
        stock_ask_buy = data.select_one("td:nth-child(8)").text
        stock_ask_sell = data.select_one("td:nth-child(9)").text
        stock_per = data.select_one("td:nth-child(11)").text
        stock_roe = data.select_one("td:nth-child(12)").text
        sheet.cell(row=i, column=1).value = int(stock_rank)
        sheet.cell(row=i, column=2).value = stock_name
        sheet.cell(row=i, column=3).value = float(stock_price.replace(',', ''))
        sheet.cell(row=i, column=4).value = float(stock_ask_buy.replace(',', ''))
        sheet.cell(row=i, column=5).value = float(stock_ask_sell.replace(',', ''))
        sheet.cell(row=i, column=6).value = stock_per
        if stock_roe == 'N/A':
            sheet.cell(row=i, column=7).value = stock_roe
        else:
            sheet.cell(row=i, column=7).value = float(stock_roe)
        #print(stock_rank, stock_name, stock_price, stock_quant, stock_ask_buy, stock_ask_sell, stock_per, stock_roe)
    except AttributeError:
        i = i - 1
        continue

wb.save('D:/Works/PycharmProjects/001_TRAINING/015_Finance/Daum/daum_finance_20211010.xlsx')
wb.close()