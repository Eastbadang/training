import openpyxl

wb = openpyxl.Workbook()
sheet1 = wb['Sheet']
sheet1.title = '수집 데이터'
sheet1['A1'] = '첫번째 시트'

sheet2 = wb.create_sheet('정리 결과')
sheet2.cell(row=1, column=1).value = '두번째 시트'

sheet1.append(['다시', '첫번째 시트'])

wb.save('test3.xlsx')